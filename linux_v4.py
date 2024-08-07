"""
室内407定位算法
估计距离比的算法函数
传入必要参数，得到估计距离比的一元三次方程
x,y坐标可以使用,对于伪逆阵标识(精度可能会差)
增加了dbm转w的公式，增加了rsrp-157这一块
"""
import numpy as np
from openpyxl import load_workbook
import sympy as sp
import math
import socket
import json

# file = load_workbook("D:/database/室内数据/data_processed/data_processed/data_all/2_2.xlsx")
# file = load_workbook("D:/database/OPPO/intel_4.xlsx")
file = load_workbook("世炬_工参.xlsx")
ws = file['Sheet1']

# [ [serv_rsrp, serv_long, serv_la, unit(1->x,y 2->gps)]
#   [nbr1_rsrp, nbr1_long, nbr2_la, unit(1->x,y 2->gps)]
#   [nbr2_rsrp, nbr2_long, nbr2_la, unit(1->x,y 2->gps)]
#   .....] data float 二维列表 list


# len(data) - 2 nbr


# {77: -73.94, 80: -72.25, 79: -69.25, 82: -68.56}
def indoor_inf(data_):

    list_positions = []
    list_positions_MAP = []
    list_err = []
    list_err_MAP = []
    num_ = 2
    len_go = len(data_)
    # sdran 命名空间，不同的pod
    # 名称是什么，
    data_list = list(map(lambda g: g[1], data_.items()))
    '''接收到的数据字典中的rsrp条目（个数应该是4-6个）'''
    pci_list = list(data_.keys())
    '''接收到的数据字典中的pci条目（个数应该是4-6个）'''
    # row = pci_list.pop()
    len_data = len(data_list)
    rsrp_list_, x, y, c = [], [], [], 1
    for i in range(len_data):
        rsrp_list_.append(float(data_list[i]))
    rsrp_list = data_process_dbm_w(rsrp_list_)
    for i in range(len(pci_list)):
        max_row = find_pci_localization()
        for j in range(2, max_row + 1):
            if pci_list[i] == ws.cell(row=j, column=1).value:
                x_, y_ = ws.cell(row=j, column=2).value, ws.cell(row=j, column=3).value
                x.append(x_), y.append(y_)
                '''x，y是两个列表，对应位置元素组成经纬度坐标'''
                x_, y_ = 0, 0
                break
    create_pci, create_rsrp, create_x, create_y, pci_var, rsrp_var, x_var, y_var = \
        locals(), locals(), locals(), locals(), [], [], [], []
    for i, var_intel in enumerate(pci_list):
        create_pci['pci_' + str(i)] = var_intel
        pci_var.append('pci_' + str(i))
    for i, var_intel in enumerate(rsrp_list):
        create_rsrp['rsrp_' + str(i)] = var_intel
        rsrp_var.append('rsrp_' + str(i))
    for i, var_intel in enumerate(x):
        create_x['latitude_' + str(i)] = var_intel
        x_var.append('latitude_' + str(i))
    for i, var_intel in enumerate(y):
        create_y['longitude_' + str(i)] = var_intel
        y_var.append('longitude_' + str(i))
    '''用键值对的形式存储了每一个接收到的pci的相关内容。create_xxx是字典形式，xxx_var是列表'''
    # 能否一并传进来基站坐标捏
    # print("第", data[0], "行数据")
    list_ = []
    temp = 0

    # print("开始解算第", int(row), "条数据的坐标：")

    for i in range(1, len(rsrp_list)):
        list_1, temp = estimated_distance_ratio(rsrp_list[i], rsrp_list[0], temp)
        '''temp是估计距离比个数，应该是pci个数减一'''
        list_.append(list_1)
        '''一元三次方程的解组成的列表，每一个方程的解当成一个元素放到list_列表中'''
    delta__, test_1, len_1 = {}, {}, {}

    for i in range(0, temp):
        for j in range(0, 3):
            if type(list_[i][j]) is sp.core.mul.Mul:
                com = isinstance(list_[i][j], sp.core.mul.Mul)
                if com is not True:
                    test_1[j] = np.e ** list_[i][j]
            elif type(list_[i][j]) is sp.core.add.Add:
                com = isinstance(list_[i][j], sp.core.add.Add)
                if com is not True:
                    test_1[j] = np.e ** list_[i][j]
            else:
                test_1[j] = np.e ** list_[i][j]
            j += 1
            """这个j+=1是不是多余了，for循环会自动加一
            test_1这个列表，包含了所有的实数根，虚数根剔除，由于求出的形式是log形式的，所以要取e的次幂"""
        delta__[i] = MR_delta_star(test_1)
        # 最佳的根选的哪一个，放到delta__列表
        if delta__[i] is False:
            break
            # 没有找到实数根
        len_1.clear()
        i += 1

    for i in range(temp):
        if delta__[i] == None:
            print('数据有问题，解算距离比异常，此条数据无法结算。该数据是第：', row_i, '条数据')
            return

    A, B, c = [], [], 0

    for i in range(1, len(pci_var)):
        A_, B_, c = A_B(x[0], y[0], x[i], y[i], c, delta__)
        A.append(A_), B.append(B_)
    '''A,B是公式13中的A和b矩阵，A是系数矩阵，b是观测矩阵，b中的参数都是已知的，当让delta__是求出来的'''

    A_temp = A
    a = np.transpose(A_temp)
    '''a是A的转置'''
    v_temp = np.dot(a, A)
    v_b = v_temp.astype(np.float64)
    try:
        v_left = np.linalg.inv(v_b)
    except np.linalg.LinAlgError:
        v_left = np.linalg.pinv(v_b)
        print('求解的矩阵是伪逆阵，误差可能会变大')
    '''求逆矩阵，但是这个函数要求矩阵必须是可逆阵'''
    v_res = np.dot(v_left, a)
    v_star = np.dot(v_res, B)
    '''最小二乘结束，求出列向量v_start = [x，y，x^2+y^2]T，这个v_start是估计的值，并不是真实的，或者说这是一个很粗糙的定位结果'''
    D = [[1, 0], [0, 1], [1, 1]]
    D_temp = D
    d = np.transpose(D_temp)
    D_ = np.dot(d, D)
    d_ = D_.astype(np.float64)
    f_left = np.linalg.inv(d_)
    f_temp = np.dot(f_left, d)
    s_star_right = [v_star[0] ** 2, v_star[1] ** 2, v_star[2]]
    s_star_res = np.dot(f_temp, s_star_right)
    if s_star_res[0] < 0:
        x_res = (-s_star_res[0]) ** 0.5
        s_star_x = x_res
    else:
        x_res = s_star_res[0] ** 0.5
        s_star_x = np.sign(v_star[0]) * x_res
    if s_star_res[1] < 0:
        y_res = (-s_star_res[1]) ** 0.5
        s_star_y = y_res
    else:
        y_res = s_star_res[1] ** 0.5
        s_star_y = np.sign(v_star[1]) * y_res
    x_float = s_star_x[0]
    y_float = s_star_y[0]
    result_list = []
    '''两部最小二乘得到的初步的定位坐标'''
    # x_res, y_res, err = change_x_y(x_float, y_float, ws.cell(row=num_, column=1).value, ws.cell(row=num_, column=2).value)
    print("LLS Latitude is: ", x_res)
    print("LLS Longitude is: ", y_res)
    result_list.append(x_res)
    result_list.append(y_res)
    # print("LLS Error is: ", err)
    # 给出误差计算公式
    psi = MAP_LM(x_res, y_res, x, y, 3.5, num_, 0.95, delta__, 3.9, pci_list, rsrp_list)
    print("MAP Latitude is: ", psi[0])
    print("MAP Longitude is: ", psi[1])


    result_list.append(psi) # x y


    # print("MAP Error is: ", psi[2])
    list_positions.append([x_res, y_res])
    list_positions_MAP.append([float(psi[0]), float(psi[1])])

    list_.clear()
    A.clear()
    B.clear()
    s_star_right.clear()
    len_1.clear()

def find_pci_localization():
    len_shape = int(ws.max_row)

    def get_max_row(wb, test):
        real_max_row = 0
        while test > 0:
            row_dict = {test.value for test in wb[test]}
            if row_dict == {None}:
                test = test - 1
            else:
                real_max_row = test
                break
        return real_max_row

    max_row = get_max_row(ws, len_shape)
    return int(max_row)

"""
# 这里是find_pci_localization函数后续需要添加的内容，隐于函数中的，不影响目前计算，后续优化需要添加
for i in range(max_row):
    if pci == ws.cell(row=i, column=1).value:
        # 此处返回坐标
        return float(ws.cell(row=i, column=2).value), float(ws.cell(row=i, column=3).value)
"""

def data_process_db_w(rsrp_list):
    list_ = []
    for i in range(len(rsrp_list)):
        list_.append(10 ** ((rsrp_list[i]) / 10))
    return list_


def data_process_dbm_w(rsrp_list):
    list_ = []
    for i in range(len(rsrp_list)):
        list_.append(10 ** ((rsrp_list[i] - 30) / 10))
    return list_


def estimated_distance_ratio(rsrp, rsrp_, temp):
    sigma, miu_n = 0.95, 3.5
    sigma_m = 3.981
    test_log_delta_ = sp.symbols('test_log_delta_')
    lambda_1_square = (0.1 * math.log(10)) ** 2 * (sigma_m ** 2) * 2
    q_1 = sp.log(rsrp / rsrp_)
    func_1 = lambda_1_square * sigma ** 4 * test_log_delta_ ** 3 - miu_n * sigma * q_1 * lambda_1_square * test_log_delta_ ** 2 + (
            (sigma ** 2 + miu_n ** 2) * lambda_1_square ** 2 - sigma ** 2 * q_1 ** 2 * lambda_1_square) \
             * test_log_delta_ + miu_n * q_1 * lambda_1_square ** 2
    # solved_delta_1 = sp.solveset(func_1, test_log_delta_)
    solved_delta_1 = sp.solve(func_1, test_log_delta_)
    list_ = list(solved_delta_1)
    temp = temp + 1
    return list_, temp

def A_B(x, y, x_, y_, c, delta__):
    A_1 = [2 * (x_ - x * delta__[c] ** 2), 2 * (y_ - y * delta__[c] ** 2), delta__[c] ** 2 - 1]
    B_1 = [x_ ** 2 + y_ ** 2 - delta__[c] ** 2 * (x ** 2 + y ** 2)]
    c = c + 1
    return A_1, B_1, c

def indoor_demo():
    # pci, rsrp, pci_11, rsrp_1, pci_21, rsrp_2, pci_31, rsrp_3 = 77, -61.88, 79, -75.25, 82, -63.13, 80, -77.38
    pci, rsrp, pci_11, rsrp_1, pci_21, rsrp_2, pci_31, rsrp_3 = 80, -77.38, 77, -61.88, 79, -75.25, 82, -63.13
    # pci, rsrp, pci_11, rsrp_1, pci_21, rsrp_2, pci_31, rsrp_3 = 79, -53, 77, -76.13, 80, -75.63, 82, -68.25
    miu_n = 3.5
    sigma_m, sigma_n = 0.95, 0.95
    list_ = []
    test_log_delta_ = sp.symbols('test_log_delta_')
    lambda_1_square = (0.1 * math.log(10)) ** 2 * (sigma_m ** 2) * 2
    q_1, q_2, q_3 = sp.log(rsrp_1 / rsrp), sp.log(rsrp_2 / rsrp), sp.log(rsrp_3 / rsrp)
    func_1 = lambda_1_square * sigma_n ** 4 * test_log_delta_ ** 3 - miu_n * sigma_n * q_1 * lambda_1_square * test_log_delta_ ** 2 + (
            (sigma_n ** 2 + miu_n ** 2) * lambda_1_square ** 2 - sigma_n ** 2 * q_1 ** 2 * lambda_1_square) \
             * test_log_delta_ + miu_n * q_1 * lambda_1_square ** 2
    solved_delta_1 = sp.solve(func_1, test_log_delta_)
    list_1 = list(solved_delta_1)
    list_.append(list_1)
    func_2 = lambda_1_square * sigma_n ** 4 * test_log_delta_ ** 3 - miu_n * sigma_n * q_2 * lambda_1_square * test_log_delta_ ** 2 + (
            (sigma_n ** 2 + miu_n ** 2) * lambda_1_square ** 2 - sigma_n ** 2 * q_2 ** 2 * lambda_1_square) \
             * test_log_delta_ + miu_n * q_2 * lambda_1_square ** 2
    # solved_delta_2 = sp.solveset(func_2, test_log_delta_)
    solved_delta_2 = sp.solve(func_2, test_log_delta_)
    list_2 = list(solved_delta_2)
    list_.append(list_2)
    func_3 = lambda_1_square * sigma_n ** 4 * test_log_delta_ ** 3 - miu_n * sigma_n * q_3 * lambda_1_square * test_log_delta_ ** 2 + (
            (sigma_n ** 2 + miu_n ** 2) * lambda_1_square ** 2 - sigma_n ** 2 * q_3 ** 2 * lambda_1_square) \
             * test_log_delta_ + miu_n * q_3 * lambda_1_square ** 2
    # solved_delta_3 = sp.solveset(func_3, test_log_delta_)
    solved_delta_3 = sp.solve(func_3, test_log_delta_)
    list_3 = list(solved_delta_3)
    list_.append(list_3)
    a_1 = np.e ** list_[0][0]
    a_2 = np.e ** list_[1][0]
    a_3 = np.e ** list_[2][0]
    A, B = [], []

    # 77: (7.739, 4.446); 79: (1, 1); 80: (1.683, 6.232); 82: (7.739, 1.561)

    A_1 = [2 * (7.739 - 1.683 * a_1 ** 2), 2 * (4.446 - 6.232 * a_1 ** 2), a_1 ** 2 - 1]
    A.append(A_1)
    B_1 = [7.739 ** 2 + 4.446 ** 2 - a_1 ** 2 * (1.683 ** 2 + 6.232 ** 2)]
    B.append(B_1)

    A_2 = [2 * (1 - 1.683 * a_2 ** 2), 2 * (1 - 6.232 * a_2 ** 2), a_2 ** 2 - 1]
    A.append(A_2)
    B_2 = [1 ** 2 + 1 ** 2 - a_2 ** 2 * (1.683 ** 2 + 6.232 ** 2)]
    B.append(B_2)

    A_3 = [2 * (7.739 - 1.683 * a_3 ** 2), 2 * (1.561 - 6.232 * a_3 ** 2), a_3 ** 2 - 1]
    A.append(A_3)
    B_3 = [7.739 ** 2 + 1.561 ** 2 - a_3 ** 2 * (1.683 ** 2 + 6.232 ** 2)]
    B.append(B_3)

    A_temp = A
    a = np.transpose(A_temp)
    v_temp = np.dot(a, A)
    v_b = v_temp.astype(np.float64)
    # Matrix_Debug.Adjugate_matrix(v_b)
    v_left = np.linalg.inv(v_b)
    # debug_res = Matrix_Debug.matrix_Debug(v_b, v_left)
    # print(debug_res)
    v_res = np.dot(v_left, a)
    v_star = np.dot(v_res, B)
    D = [[1, 0], [0, 1], [1, 1]]
    D_temp = D
    d = np.transpose(D_temp)
    D_ = np.dot(d, D)
    d_ = D_.astype(np.float64)
    f_left = np.linalg.inv(d_)
    f_temp = np.dot(f_left, d)
    s_star_right = [v_star[0] ** 2, v_star[1] ** 2, v_star[2]]
    s_star_res = np.dot(f_temp, s_star_right)
    print(s_star_res)

def change_pci(p, x, y, r, p_, x_, y_, r_):
    if r >= r_:
        p, x, y, r, p_, x_, y_, r_ = p_, x_, y_, r_, p, x, y, r
        return p, x, y, r, p_, x_, y_, r_
    else:
        return p, x, y, r, p_, x_, y_, r_

def change_x_y(x, y, x_true, y_true):
    err_0 = ((x - x_true) ** 2 + (y - y_true) ** 2) ** 0.5
    err_1 = ((y - x_true) ** 2 + (x - y_true) ** 2) ** 0.5
    if err_0 <= err_1:
        return x, y, err_0
    else:
        return y, x, err_1

def MAP_LM(latitude, longitude, latitude_list, longitude_list, miu_n, num, sigma_gamma, delta, n, pci_list, rsrp_list):
    # err = ((latitude - ws.cell(row=num, column=1).value) ** 2 + (longitude - ws.cell(row=num, column=2).value) ** 2) ** 0.5
    # print("LLS误差为：", err)
    list_res = []
    list_err_res, list_ssd = [], []
    miu = 0
    v = 2
    err__ = []
    to = 10 ** (-3)
    for i in range(len(pci_list)):
        list_ssd.append(6)
    psi = np.matrix([latitude, longitude, n]).T

    psi_n = psi
    h_ = np.matrix([0, 0, 1 / sigma_gamma]).T
    h = np.dot(h_, np.transpose(h_))
    I = [[1, 0, 0],
         [0, 1, 0],
         [0, 0, 1]]
    for k in range(1, 15):
        javobian_matrix, q, delta_, drss = \
        jacob_matrix_func(longitude, latitude, latitude_list, longitude_list, float(psi[2]), delta, list_ssd, rsrp_list)
        j_transpose = np.transpose(javobian_matrix)
        q_ = q.astype(np.float64)
        q_inv = np.linalg.inv(q_)
        j_q = np.dot(j_transpose, q_inv)
        j_q_j_t = np.dot(j_q, javobian_matrix)
        if k == 1:
            miu = to * np.max(j_q_j_t)
        a = j_q_j_t + h + np.multiply(miu, I)
        a_ = a.astype(np.float64)
        a_inv = np.linalg.inv(a_)

        j_q_j_t_j = np.dot(a_inv, j_transpose)
        j_q_j_t_j_q = np.dot(j_q_j_t_j, q_inv)
        delta_drss = drss + np.multiply(float(psi_n[2]), delta_)
        delta_drss_t = np.transpose(delta_drss)
        left = np.dot(j_q_j_t_j_q, delta_drss)
        right = np.multiply((float(psi_n[2]) - miu_n) / sigma_gamma, h_)
        res = left + right

        h_lm_1 = j_q_j_t + np.multiply(miu, I)
        h_lm_1_ = h_lm_1.astype(np.float64)
        h_lm_1_inv = np.linalg.inv(h_lm_1_)
        h_lm_2 = np.dot(h_lm_1_inv, j_transpose)
        h_lm_3 = np.dot(h_lm_2, q_inv)
        h_lm = np.dot(h_lm_3, delta_drss)
        h_lm_t = np.transpose(h_lm)

        numerator_1 = np.multiply(2, delta_drss_t)
        numerator_2 = np.dot(numerator_1, q_inv)
        numerator_3 = np.dot(numerator_2, javobian_matrix)
        numerator_4 = np.dot(numerator_3, h_lm)

        numerator_5 = np.dot(h_lm_t, j_transpose)
        numerator_6 = np.dot(numerator_5, q_inv)
        numerator_7 = np.dot(numerator_6, javobian_matrix)
        numerator_8 = np.dot(numerator_7, h_lm)

        numerator = numerator_4 - numerator_8

        denominator_1 = np.multiply(miu, h_lm)
        denominator_2 = np.dot(j_q, delta_drss)
        denominator_3 = denominator_1 + denominator_2

        denominator = np.dot(h_lm_t, denominator_3)

        rou = float(numerator / denominator)


        psi_res = lm_psi_test(psi, res)
        # psi_n = psi_res
        list_add_psi_err = [float(psi_res[0]), float(psi_res[1])]
        list_res.append(list_add_psi_err)
        # list_add_psi_err.clear()

        if rou > 2:
            miu = miu * max(1 / 3, 1 - (2 * rou - 1) ** 3)
            v = 2
            psi_n = [float(psi_n[0]) - float(res[1]), float(psi_n[1]) - float(res[0]), float(psi_n[2]) - float(res[2])]
            err_ = ((psi_n[0] - ws.cell(row=num, column=1).value) ** 2 + (psi_n[1] - ws.cell(row=num, column=2).value) ** 2) ** 0.5
            err__.append(err_)
            return psi_n, err_
        else:
            miu = miu * v
            v = 2 * v

        if max(np.dot(j_q, delta_drss)) > 10 ** 10:
            return psi_n
    
        if k == 14:
            # for i in range(len(list_res)):
            #     if list_res[i] == list_res[i + 1]:
            #         return list_res[i - 1]
            #     elif i == 13:
            return list_res[0]

def jacob_matrix_func(longitude, latitude, latitude_list, longitude_list, miu_n, delta, ssd, rsrp_list):
    drss_, delta_, delta_temp = [], [], []
    for i in range(1, len(rsrp_list)):
        drss_.append(rsrp_list[i] - rsrp_list[0])
    drss = np.matrix(drss_).T
    for i in range(len(delta)):
        delta_temp.append(sp.log(delta[i]))
    delta_ = np.matrix(delta_temp).T
    I_left = np.matrix([-1, -1, -1]).T
    I = np.append(I_left, np.eye(len(ssd) - 1), axis=1)
    I_transpose = np.transpose(I)
    q_matrix_ = np.diag([6.3] * len(ssd))
    q_matrix_temp = np.dot(I, q_matrix_)
    q_matrix = np.dot(q_matrix_temp, I_transpose)
    jacobian_matrix, jacobian_temp = [], []
    for i in range(1, len(rsrp_list)):
        jacobian_temp = [(miu_n * (latitude - latitude_list[i]) / ((latitude - latitude_list[i]) ** 2 +
                                                                 (longitude - longitude_list[i]) ** 2)) -
                             (miu_n * (latitude - latitude_list[0]) / ((latitude - latitude_list[0]) ** 2 +
                                                                 (longitude - longitude_list[0]) ** 2)),
                         (miu_n * (longitude - longitude_list[i]) / ((latitude - latitude_list[i]) ** 2 +
                                                                     (longitude - longitude_list[i]) ** 2)) -
                         (miu_n * (longitude - longitude_list[0]) / ((latitude - latitude_list[0]) ** 2 +
                                                                     (longitude - longitude_list[0]) ** 2)),
                         delta[i - 1]]
        jacobian_matrix.append(jacobian_temp)

    return jacobian_matrix, q_matrix, delta_, drss

def lm_psi(psi_n, res_, long, lat):

    psi_n_1 = [float(psi_n[0]) - abs(float(res_[1])), float(psi_n[1]) - abs(float(res_[0])),
                 float(psi_n[2]) + float(res_[2])]
    err_1 = ((psi_n_1[0] - lat) ** 2 + (psi_n_1[1] - long) ** 2) ** 0.5

    psi_n_2 = [float(psi_n[0]) - abs(float(res_[1])), float(psi_n[1]) + abs(float(res_[0])),
               float(psi_n[2]) + float(res_[2])]
    err_2 = ((psi_n_2[0] - lat) ** 2 + (psi_n_2[1] - long) ** 2) ** 0.5

    psi_n_3 = [float(psi_n[0]) + abs(float(res_[1])), float(psi_n[1]) - abs(float(res_[0])),
               float(psi_n[2]) + float(res_[2])]
    err_3 = ((psi_n_3[0] - lat) ** 2 + (psi_n_3[1] - long) ** 2) ** 0.5

    psi_n_4 = [float(psi_n[0]) + abs(float(res_[1])), float(psi_n[1]) + abs(float(res_[0])),
               float(psi_n[2]) + float(res_[2])]
    err_4 = ((psi_n_4[0] - lat) ** 2 + (psi_n_4[1] - long) ** 2) ** 0.5

    psi_n_1_ = [float(psi_n[0]) - abs(float(res_[0])), float(psi_n[1]) - abs(float(res_[1])),
                 float(psi_n[2]) + float(res_[2])]
    err_1_ = ((psi_n_1_[0] - lat) ** 2 + (psi_n_1_[1] - long) ** 2) ** 0.5

    psi_n_2_ = [float(psi_n[0]) - abs(float(res_[0])), float(psi_n[1]) + abs(float(res_[1])),
               float(psi_n[2]) + float(res_[2])]
    err_2_ = ((psi_n_2_[0] - lat) ** 2 + (psi_n_2_[1] - long) ** 2) ** 0.5

    psi_n_3_ = [float(psi_n[0]) + abs(float(res_[0])), float(psi_n[1]) - abs(float(res_[1])),
               float(psi_n[2]) + float(res_[2])]
    err_3_ = ((psi_n_3_[0] - lat) ** 2 + (psi_n_3_[1] - long) ** 2) ** 0.5

    psi_n_4_ = [float(psi_n[0]) + abs(float(res_[0])), float(psi_n[1]) + abs(float(res_[1])),
               float(psi_n[2]) + float(res_[2])]
    err_4_ = ((psi_n_4_[0] - lat) ** 2 + (psi_n_4_[1] - long) ** 2) ** 0.5


    list_err = [float(err_1), float(err_2), float(err_3), float(err_4), float(err_1_), float(err_2_), float(err_3_), float(err_4_)]
    # -, -
    if err_1 == min(list_err):
        return psi_n_1, err_1
    # -, +
    elif err_2 == min(list_err):
        return psi_n_2, err_2
    # +, -
    elif err_3 == min(list_err):
        return psi_n_3, err_3
    # +, +
    elif err_4 == min(list_err):
        return psi_n_4, err_4
    elif err_1_ == min(list_err):
        return psi_n_1_, err_1_
    elif err_2_ == min(list_err):
        return psi_n_2_, err_2_
    elif err_3_ == min(list_err):
        return psi_n_3_, err_3_
    elif err_4_ == min(list_err):
        return psi_n_4_, err_4_

def lm_psi_test(psi_n, res_):
    psi_n_1 = [float(psi_n[0]) - abs(float(res_[1])), float(psi_n[1]) + abs(float(res_[0])),
                 float(psi_n[2]) + float(res_[2])]
    return psi_n_1


def MR_delta_star(list_):
    # list_star = {}
    for i in range(0, len(list_)):
        if len(list_) == 1:
            delta_ = list_[0]
            return delta_
        elif len(list_) == 0 or len(list_) > 3:
            print("传入list有误，是个空列表，应当检查列表添加部分")
            return False
        else:
            # 这里如果出现不止一个数值，有两种做法：
            # 一般不会出现很高的距离比，所以可以找最小值，
            # 或者找靠近1的值
            if len(list_) == 2 or len(list_) == 3:
                delta_ = min(list_)
                return delta_


IP = "localhost"
port = 40005
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
s.bind((IP, port))
s.listen(5)
print('listen at port :', port)
conn, addr = s.accept()
# addr = '10.245.24.98'
print('connected by', addr)
while True:
    try:
        data = conn.recv(1024)
        data = data.hex()
        # data = data.decode()
        '''开始截取数据'''
        row, ID, pci, rsrp = [], [], [], []
        row_i = data[0:16]
        ID_i = data[16:32]
        pci_0 = data[32:36]
        rsrp_0 = data[36:40]
        pci_1 = data[40:44]
        rsrp_1 = data[44:48]
        pci_2 = data[48:52]
        rsrp_2 = data[52:56]
        pci_3 = data[56:60]
        rsrp_3 = data[60:64]
        if row_i == '0000000000000000' or ID_i == '0000000000000000' or pci_0 == '00' or rsrp_0 == '00' or \
                pci_1 == '00' or rsrp_1 == '00' or pci_2 == '00' or rsrp_2 == '00' or pci_3 == '00' or rsrp_3 == '00':
            print('传入数据有误，可能是pci或rsrp为空')
            break
        pci.append(int(pci_0, 16))
        rsrp.append(int(rsrp_0, 16) - 157)
        pci.append(int(pci_1, 16))
        rsrp.append(int(rsrp_1, 16) - 157)
        pci.append(int(pci_2, 16))
        rsrp.append(int(rsrp_2, 16) - 157)
        pci.append(int(pci_3, 16))
        rsrp.append(int(rsrp_3, 16) - 157)
        data = dict(zip(pci, rsrp))
        row_i = int(row_i, 16)
        ID_i = int(ID_i, 16)
        conn.send(str.encode('1'))
        print('解算的第：', row_i, '条数据', '小区ID是', ID_i)



        indoor_inf(data, )

    except:
        conn.close()
        s.close()
        break